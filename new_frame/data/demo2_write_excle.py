# -*- coding=utf-8 -*-
# Author: BoLin Chen
# @Date : 2019-08-08

import openpyxl
import time


data = openpyxl.load_workbook('FDD接口测试用例.xlsx')
sheets = data.get_sheet_names()  # 获得所有xlsx的页签名称列表

sheet_ddsf = data.get_sheet_by_name("多多商服")  # 打开指定名称的页签
rows = sheet_ddsf.rows   # 获得行数据，类型为迭代器
cols = sheet_ddsf.columns  # 获得列数据，类型为迭代器
# print(dir(sheet_ddsf))

# num_rows = len(list(rows))
# num_cols = len(list(cols))
# print(num_rows, num_cols)

num_rows = sheet_ddsf.max_row
num_cols = sheet_ddsf.max_column
# print(num_rows, num_cols)
#
for i in range(2, num_rows+1):
	sheet_ddsf.cell(i, num_cols).value = 'fail'
	# print(sheet_ddsf.cell(i, num_cols).value)

data.save('FDD接口测试用例.xlsx')




# for row in rows:
# 	# print(row)
# 	line = [col.value for col in row]
# 	print(line)

# print(sheet_ddsf.cell(1, 1))